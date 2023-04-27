import openpyxl


class TestData:
    # no need to create object from this class to call that method, because it's staticmethod TestData.getTestData
    @staticmethod
    def getTestData(test_case_name):
        data_dict = {}
        book = openpyxl.load_workbook("F:\\Projects\\PythonProjects\\SelAutomationTesting\\testsData\\TestsData.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[data_dict]  # return as a list so fixture can transfer the data


